"""Экспорт результатов в Excel (.xlsx).

Структура книги:
  Лист «Сравнение»   — сводная таблица КПЭ по всем сценариям; ячейки окрашены
                        по статусу (зелёный / жёлтый / красный).
  Лист «Аудит»       — длинная таблица: все поля × все сценарии со source/formula.
  Лист «Сценарий N»  — (опционально) краткая сводка каждого сценария.

Использование:
    from urban_model.export import to_xlsx
    to_xlsx(pairs, "output/tep_report.xlsx")
"""

from __future__ import annotations

import pathlib
import re
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from urban_model.export.table import results_to_audit_dataframe, results_to_dataframe
from urban_model.models.result import TEPResult

# ---------------------------------------------------------------------------
# Палитра статусов (Excel ColorIndex-совместимые HEX без #)
# ---------------------------------------------------------------------------
_FILL = {
    "ok":       PatternFill("solid", fgColor="C6EFCE"),   # светло-зелёный
    "warning":  PatternFill("solid", fgColor="FFEB9C"),   # светло-жёлтый
    "error":    PatternFill("solid", fgColor="FFC7CE"),   # светло-красный
    "manual":   PatternFill("solid", fgColor="BDD7EE"),   # светло-синий
    "no_data":  PatternFill("solid", fgColor="E2EFDA"),   # очень светлый
    "ДЕФИЦИТ":  PatternFill("solid", fgColor="FFC7CE"),
    "OK":       PatternFill("solid", fgColor="C6EFCE"),
}
_FILL_HEADER = PatternFill("solid", fgColor="2F5496")   # тёмно-синий
_FILL_LABEL  = PatternFill("solid", fgColor="D6DCE4")   # серый

_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_LABEL  = Font(bold=True, size=10)
_FONT_BODY   = Font(size=10)

_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STATUS_ROWS = {"Плотность статус", "Статус баланса территории"}

# Серый фон строк-заголовков секций сравнительной таблицы (v0.16.2).
_FILL_SECTION = PatternFill("solid", fgColor="D9DCE0")
_FONT_SECTION = Font(size=10, bold=True)


def _auto_width(ws, min_col: int = 1, min_width: int = 12, max_width: int = 60) -> None:
    """Подогнать ширину столбцов по содержимому."""
    for col_cells in ws.iter_cols():
        col_idx = col_cells[0].column
        if col_idx < min_col:
            continue
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(max_len + 2, max_width)
        )


def _write_comparison_sheet(wb: Workbook, pairs: list[tuple[str, TEPResult]]) -> None:
    ws = wb.active
    ws.title = "Сравнение"

    df = results_to_dataframe(pairs)
    scenario_names = [name for name, _ in pairs]

    # --- Заголовок ---
    ws.append(["Показатель"] + scenario_names)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 28

    # --- Данные ---
    # Соберём статусы по сценарию для каждой строки индекса
    status_map: dict[str, list[str]] = {
        "Плотность статус": [
            res.density_chel_per_ga.status.value for _, res in pairs
        ],
        "Статус баланса территории": [
            "OK" if res.balance.is_feasible else "ДЕФИЦИТ" for _, res in pairs
        ],
    }

    from urban_model.export.table import KPI_SECTION_LABELS

    for row_idx, (label, row_data) in enumerate(df.iterrows(), start=2):
        # v0.16.2 (п.4): строка-заголовок секции — серый фон по всей ширине.
        is_section = str(label) in KPI_SECTION_LABELS
        lab_cell = ws.cell(row=row_idx, column=1, value=label)
        lab_cell.fill = _FILL_SECTION if is_section else _FILL_LABEL
        lab_cell.font = _FONT_SECTION if is_section else _FONT_LABEL
        lab_cell.border = _BORDER

        statuses = status_map.get(str(label), [None] * len(scenario_names))

        for col_idx, (val, status_str) in enumerate(zip(row_data, statuses), start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if is_section:
                cell.fill = _FILL_SECTION
            # Окрашиваем строки-статусов
            elif str(label) in _STATUS_ROWS and status_str in _FILL:
                cell.fill = _FILL[status_str]
            # Для остальных строк — только значение

    _auto_width(ws)
    ws.freeze_panes = "B2"


def _write_audit_sheet(wb: Workbook, pairs: list[tuple[str, TEPResult]]) -> None:
    ws = wb.create_sheet("Аудит")
    df = results_to_audit_dataframe(pairs)
    if df.empty:
        return

    # Заголовок
    cols = list(df.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for _, row in df.iterrows():
        ws.append(list(row))
        data_row = ws[ws.max_row]
        status_val = str(row.get("статус", "")).lower()
        fill = _FILL.get(status_val)
        for cell in data_row:
            cell.font = _FONT_BODY
            cell.border = _BORDER
            if fill and cell.column_letter != "A":
                cell.fill = fill

    _auto_width(ws)
    ws.freeze_panes = "A2"


def to_xlsx(
    pairs: list[tuple[str, TEPResult]],
    path: Union[str, pathlib.Path],
) -> pathlib.Path:
    """Записать результаты расчётов в Excel-файл.

    Args:
        pairs: список (название_сценария, TEPResult). Порядок определяет порядок
               столбцов в листе «Сравнение» и строк в листе «Аудит».
        path:  Путь к выходному файлу (.xlsx). Директория должна существовать.

    Returns:
        pathlib.Path к записанному файлу.

    Example:
        >>> from urban_model.export import to_xlsx
        >>> to_xlsx([("Малый", res1), ("Большой", res2)], "report.xlsx")
    """
    path = pathlib.Path(path)
    wb = Workbook()

    _write_comparison_sheet(wb, pairs)
    _write_audit_sheet(wb, pairs)

    wb.save(path)
    return path


# ===========================================================================
# Паспорт варианта (v0.12.25): единый комплексный xlsx на ОДИН вариант.
# Лист «Сводка» (секции по категориям) + лист «Баланс территории».
# Используется на «Расчёте» и на 4 карточках подбора.
# ===========================================================================

_FILL_SECTION = PatternFill("solid", fgColor="2F5496")   # тёмно-синий
_FONT_SECTION = Font(bold=True, color="FFFFFF", size=11)
_FILL_TITLE = PatternFill("solid", fgColor="1F3864")
_FONT_TITLE = Font(bold=True, color="FFFFFF", size=13)
_FILL_TOTAL = PatternFill("solid", fgColor="D6DCE4")
_FONT_TOTAL = Font(bold=True, size=10)

_VRI_NAMES = {
    "4.4": "Магазины (торговля)",
    "4.6": "Общественное питание",
    "3.3": "Бытовое обслуживание",
    "3.4.1": "Поликлиника",
    "3.5.1": "Доп. образование",
}
_CLASS_NAMES = {"economy": "Эконом", "comfort": "Комфорт", "business": "Бизнес"}


def _num(v, nd: int = 0) -> str:
    """Число с пробелами-разделителями тысяч; '—' для None."""
    if v is None:
        return "—"
    try:
        fv = float(v)
    except (TypeError, ValueError):
        return str(v)
    if nd == 0 and abs(fv - round(fv)) < 1e-9:
        return f"{int(round(fv)):,}".replace(",", " ")
    return f"{fv:,.{nd}f}".replace(",", " ")


def _buckets(formula: str | None) -> list[int]:
    if not formula or "[" not in formula:
        return []
    m = re.search(r"\[([^\]]+)\]", formula)
    if not m:
        return []
    try:
        return [int(x.strip()) for x in m.group(1).split(",") if x.strip()]
    except ValueError:
        return []


def _variant_sections(result: TEPResult, options) -> list[tuple[str, list[tuple[str, str]]]]:
    """Сформировать секции «Сводки»: [(заголовок, [(показатель, значение), …]), …]."""
    r = result
    secs: list[tuple[str, list[tuple[str, str]]]] = []

    # --- Основные ---
    floors_label = (
        " / ".join(str(d["floors"]) for d in r.floor_clusters_detail) + " эт."
        if r.floor_clusters_detail
        else (f"{int(options.floors)} эт." if options is not None else "—")
    )
    secs.append(("Основные показатели", [
        ("КИТ (ПЗЗ)", _num(r.kit.value, 3)),
        ("Нормативный потолок КИТ", _num(r.kit_normative_max.value, 2)),
        ("Плотность застройки (GFA/S кв.)", _num(r.block_density.value, 3)),
        ("Общая площадь зданий (GFA), м²", _num(r.gfa.value)),
        ("Площадь квартир, м²", _num(r.apartments_area.value)),
        ("Население, чел", _num(r.population.value)),
        (f"Плотность, чел/га (норма {_num(r.density_chel_per_ga.normative)})",
         _num(r.density_chel_per_ga.value, 1)),
        ("Этажность", floors_label),
    ]))
    if r.floor_clusters_detail:
        zrows = [("— зоны этажности —", "")]
        for d in r.floor_clusters_detail:
            zrows.append((
                f"   {d['label']}: {_num(d['area_m2'])} м², {d['floors']} эт.",
                f"КИТ {d['kit']:.3f}, квартиры {_num(d['apartments_area'])} м²",
            ))
        secs.append(("Кластеры этажности", zrows))

    # --- Жильё ---
    sell = (r.economy.sellable_ratio * 100) if r.economy else None
    cls = _CLASS_NAMES.get(getattr(options, "residential_class", ""), "—") if options else "—"
    secs.append(("Жильё", [
        ("Площадь квартир, м²", _num(r.apartments_area.value)),
        ("Пятно застройки, м²", _num(r.housing_footprint.value)),
        ("ЗУ жилой застройки, м²", _num(r.housing_lot_area.value)),
        ("Выход жилья (квартиры / GFA), %", _num(sell, 1) if sell is not None else "—"),
        ("Класс жилья", cls),
    ]))

    # --- ВПП ---
    bi_total = r.built_in_area.value or 0
    if bi_total > 0:
        vpp_rows = [("Общая площадь ВПП, м²", _num(bi_total))]
        bi_list = list(getattr(options, "built_in_list", []) or []) if options else []
        if options is not None and getattr(options, "built_in", None) is not None:
            bi_list = [options.built_in] + bi_list
        for b in bi_list:
            nm = b.label or _VRI_NAMES.get(b.vri_code, f"ВРИ {b.vri_code}")
            vpp_rows.append((f"   {nm} (ВРИ {b.vri_code})", f"{_num(b.area_m2)} м²"))
        vpp_rows.append(("Парковка ВПП, м/м", _num(r.built_in_parking_places.value)))
        vpp_rows.append(("Озеленение ВПП, м²", _num(r.built_in_greening_area.value)))
        secs.append(("ВПП (встроенно-пристроенные)", vpp_rows))

    # --- Соцобъекты ---
    soc: list[tuple[str, str]] = []
    kg_b = _buckets(r.kindergarten_places_accepted.formula)
    soc.append(("ДОО: требуется / принято, мест",
                f"{_num(r.kindergarten_places_required.value)} / {_num(r.kindergarten_places_accepted.value)}"))
    if kg_b:
        soc.append(("ДОО: число объектов / вместимости", f"{len(kg_b)} — {kg_b}"))
    soc.append(("ДОО: ЗУ / здание, м²",
                f"{_num(r.kindergarten_plot_area.value)} / {_num(r.kindergarten_building_area.value)}"))
    sch_b = _buckets(r.school_places_accepted.formula)
    soc.append(("СОШ: требуется / принято, мест",
                f"{_num(r.school_places_required.value)} / {_num(r.school_places_accepted.value)}"))
    if sch_b:
        soc.append(("СОШ: число объектов / вместимости", f"{len(sch_b)} — {sch_b}"))
    soc.append(("СОШ: ЗУ / здание, м²",
                f"{_num(r.school_plot_area.value)} / {_num(r.school_building_area.value)}"))
    ae = r.add_education_places_accepted.value or 0
    if ae > 0:
        place = "встроенное (ВПП)" if getattr(r, "add_education_built_in", False) else "отдельно стоящее"
        soc.append(("Доп. образование: мест / размещение", f"{_num(ae)} — {place}"))
        soc.append(("Доп. образование: ЗУ / здание, м²",
                    f"{_num(r.add_education_plot_area.value)} / {_num(r.add_education_building_area.value)}"))
    poly = int(getattr(r, "polyclinic_visits_accepted", None).value or 0) \
        if getattr(r, "polyclinic_visits_accepted", None) is not None else 0
    if poly > 0:
        pplace = "ВПП (офис врача)" if getattr(r, "polyclinic_built_in", False) else "отд. стоящая"
        soc.append(("Поликлиника: посещений / размещение", f"{_num(poly)} — {pplace}"))
        soc.append(("Поликлиника: ЗУ / здание, м²",
                    f"{_num(r.polyclinic_plot_area.value)} / {_num(r.polyclinic_building_area.value)}"))
    if (r.sport_facilities_area.value or 0) > 0:
        soc.append(("Спорт. сооружения: площадь / ЗУ, м²",
                    f"{_num(r.sport_facilities_area.value)} / {_num(r.sport_facilities_plot_area.value)}"))
    soc.append(("Парковки соцобъектов всего, м/м", _num(r.social_parking_total.value)))
    soc.append(("   в т.ч. ДОО / СОШ, м/м",
                f"{_num(r.social_parking_kindergarten.value)} / {_num(r.social_parking_school.value)}"))
    if ae > 0:
        soc.append(("   в т.ч. доп. образование, м/м", _num(r.add_education_parking_places.value)))
    if poly > 0:
        soc.append(("   в т.ч. поликлиника, м/м", _num(r.polyclinic_parking_places.value)))
    secs.append(("Социальные объекты", soc))

    # --- Парковки ---
    park = [("Всего требуется, м/м", _num(r.parking_required_places.value))]
    park.append(("Открытые: м/м / м²",
                 f"{_num(r.parking_open_places.value)} / {_num(r.parking_open_area.value)}"))
    if (r.parking_multilevel_places.value or 0) > 0:
        park.append(("Многоуровневые: м/м / объектов / м²",
                     f"{_num(r.parking_multilevel_places.value)} / "
                     f"{_num(r.parking_multilevel_objects.value)} / {_num(r.parking_multilevel_area.value)}"))
    if (r.parking_underground_places.value or 0) > 0:
        park.append(("Подземные, м/м", _num(r.parking_underground_places.value)))
    if (getattr(r, "parking_stylobate_places", None) is not None
            and (r.parking_stylobate_places.value or 0) > 0):
        park.append(("Стилобатные: м/м / дека м²",
                     f"{_num(r.parking_stylobate_places.value)} / {_num(r.parking_stylobate_area.value)}"))
    secs.append(("Парковки", park))

    # --- ЗНОП и озеленение / проезды ---
    secs.append(("ЗНОП и озеленение", [
        ("ЗНОП, м²/чел", _num(r.znop_per_person.value, 1)),
        ("ЗНОП всего, м²", _num(r.znop_area.value)),
        ("Озеленение жилья, м²", _num(r.greening_housing_area.value)),
        ("Требуемое озеленение квартала (норма), м²", _num(r.greening_quarter_required.value)),
    ]))
    secs.append(("Проезды", [
        ("Внутриквартальные, м²", _num(r.driveways_intra_quarter_area.value)),
        ("На ЗУ жилой застройки, м²", _num(r.driveways_housing_lot_area.value)),
    ]))

    # --- Инженерия ---
    eng = getattr(r, "engineering", None)
    if eng is not None and eng.objects:
        erows = []
        for o in eng.objects:
            if o.count <= 0:
                continue
            cap = f", {_num(o.capacity, 1)} {o.capacity_unit}" if o.capacity else ""
            erows.append((f"{o.label.split(' (')[0]}: кол-во{cap}",
                          f"{o.count} шт, ЗУ {_num(o.plot_total)} м²"))
        erows.append(("Итого ЗУ инженерии (в балансе), м²", _num(eng.plot_in_balance)))
        secs.append(("Инженерная инфраструктура", erows))

    # --- Экономика ---
    if r.economy is not None:
        e = r.economy
        _FUND_RU = {
            "compensated": "Город компенсирует %",
            "developer": "Полностью застройщик",
            "city": "Полностью город",
            "at_cost": "Передача по себестоимости",
        }
        _fund = getattr(options, "social_funding", "compensated") if options else "compensated"
        secs.append(("Экономика (условные баллы)", [
            ("Эконом-индекс (100 = окупаемость)", _num(e.economy_index)),
            ("Соцобъекты — финансирование", _FUND_RU.get(_fund, _fund)),
            ("Выход жилья, %", _num(e.sellable_ratio * 100, 1)),
            ("Себестоимость итого", _num(e.cost.total)),
            ("Выручка итого", _num(e.revenue.total)),
            ("Прибыль (условная)", _num(e.profit)),
            ("Маржа, %", _num(e.margin * 100, 1)),
            ("ROI, %", _num(e.roi * 100, 1)),
            ("Соц. нагрузка (нетто)", _num(e.net_social_burden)),
        ]))
    return secs


def _write_variant_summary(wb: Workbook, name: str, result: TEPResult, options) -> None:
    ws = wb.active
    ws.title = "Сводка"
    # Заголовок-вариант
    ws.append([f"Вариант: {name}", ""])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    c = ws.cell(row=1, column=1)
    c.fill, c.font = _FILL_TITLE, _FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    row = 2
    for title, items in _variant_sections(result, options):
        ws.append([title, ""])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sc = ws.cell(row=row, column=1)
        sc.fill, sc.font = _FILL_SECTION, _FONT_SECTION
        sc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1
        for label, value in items:
            ws.cell(row=row, column=1, value=label).font = _FONT_BODY
            vcell = ws.cell(row=row, column=2, value=value)
            vcell.font = _FONT_BODY
            vcell.alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=1).border = _BORDER
            vcell.border = _BORDER
            row += 1
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = "A2"


def _write_variant_balance(wb: Workbook, result: TEPResult) -> None:
    ws = wb.create_sheet("Баланс территории")
    b = result.balance
    site = b.site_area
    pretty = {
        "housing_lot": "ЗУ жилой застройки",
        "kindergarten_plot": "Участки ДОО",
        "school_plot": "Участки СОШ",
        "sport_facilities": "Спортивные сооружения",
        "social_parking_plot": "Парковки соцобъектов (ДОО/СОШ/доп.обр)",
        "add_education_plot": "Доп. образование (ЗУ)",
        "znop": "ЗНОП",
        "intra_quarter_driveways": "Внутриквартальные проезды",
        "parking_multilevel": "Многоуровневые паркинги",
        "custom_objects": "Пользовательские объекты",
        "engineering_plot": "Инженерная инфраструктура",
    }
    required_map = {
        "housing_lot": result.housing_lot_area.value,
        "kindergarten_plot": result.kindergarten_plot_area.value,
        "school_plot": result.school_plot_area.value,
        "sport_facilities": result.sport_facilities_plot_area.value,
        "social_parking_plot": result.social_parking_area.value,
        "add_education_plot": result.add_education_plot_area.value,
        "znop": result.znop_area.value,
        "intra_quarter_driveways": result.driveways_intra_quarter_area.value,
        "parking_multilevel": result.parking_multilevel_area.value,
        "engineering_plot": (result.engineering.plot_total_all if result.engineering else 0.0),
    }
    ws.append(["Компонент", "Требуется, м²", "В балансе, м²", "Доля квартала, %"])
    for cell in ws[1]:
        cell.fill, cell.font = _FILL_HEADER, _FONT_HEADER
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    for nm, val in sorted(b.components.items(), key=lambda kv: -kv[1]):
        req = required_map.get(nm, val)
        if req is None:
            req = val
        pct = val / site * 100 if site > 0 else 0
        ws.append([pretty.get(nm, nm), _num(req), _num(val), f"{pct:.1f}%"])
        for cell in ws[ws.max_row]:
            cell.font = _FONT_BODY
            cell.border = _BORDER

    def _total_row(label, val, fill=True):
        ws.append([label, "", _num(val), f"{val / site * 100:.1f}%" if site > 0 else ""])
        for cell in ws[ws.max_row]:
            cell.font = _FONT_TOTAL
            cell.border = _BORDER
            if fill:
                cell.fill = _FILL_TOTAL

    _total_row("Итого занято", b.required_total)
    _total_row("Резерв (surplus)", b.surplus)
    ws.append(["Площадь квартала", "", _num(site), "100.0%"])
    for cell in ws[ws.max_row]:
        cell.font = _FONT_TOTAL
        cell.border = _BORDER

    # Контроль озеленения
    ws.append([])
    ws.append(["Контроль озеленения 25%", "Факт, м²", "Требуется, м²", "Дефицит, м²"])
    for cell in ws[ws.max_row]:
        cell.fill, cell.font = _FILL_HEADER, _FONT_HEADER
        cell.border = _BORDER
    ws.append(["Озеленение квартала", _num(b.greening_actual), _num(b.greening_required),
               _num(b.greening_deficit)])
    for cell in ws[ws.max_row]:
        cell.font = _FONT_BODY
        cell.border = _BORDER

    ws.column_dimensions["A"].width = 42
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A2"


def build_variant_xlsx(
    name: str,
    result: TEPResult,
    options,
    path: Union[str, pathlib.Path],
) -> pathlib.Path:
    """Комплексный «паспорт варианта» в Excel (v0.12.25).

    Два листа:
      «Сводка»            — все категории ТЭП по секциям (жильё, ВПП, соц,
                            парковки, ЗНОП, проезды, инженерия, экономика).
      «Баланс территории» — компоненты баланса + резерв + контроль озеленения.

    Используется на «Расчёте» и на карточках подбора (один вариант на файл).
    """
    path = pathlib.Path(path)
    wb = Workbook()
    _write_variant_summary(wb, name, result, options)
    _write_variant_balance(wb, result)
    wb.save(path)
    return path
