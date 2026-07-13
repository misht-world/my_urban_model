"""Тесты v0.2: verify_kit, compare_scenarios, run_scenarios, xlsx-экспорт, Scenario."""

from __future__ import annotations

import pathlib
import tempfile

import pytest

from urban_model import compare_scenarios, run_scenarios, solve_max_kit, verify_kit
from urban_model.export import results_to_dataframe, to_xlsx
from urban_model.models import CalculationOptions, Scenario, Site
from urban_model.normatives import load_normatives


@pytest.fixture(scope="module")
def spb():
    return load_normatives("spb")


@pytest.fixture(scope="module")
def site_medium():
    return Site(area_m2=50_000, name="Средний")


# ---------------------------------------------------------------------------
# verify_kit
# ---------------------------------------------------------------------------

class TestVerifyKit:
    def test_returns_tepresult(self, spb, site_medium):
        """verify_kit принимает block_density (внутреннюю переменную бисекции).
        result.block_density == входному значению; result.kit — это КИТ ПЗЗ."""
        res = verify_kit(1.8, site_medium, CalculationOptions(floors=12), spb)
        assert res.block_density.value == 1.8

    def test_source_present(self, spb, site_medium):
        res = verify_kit(1.5, site_medium, norms=spb)
        assert res.population.source is not None

    def test_status_ok_for_low_kit(self, spb):
        """block_density=0.5 на большом квартале — плотность в норме."""
        res = verify_kit(0.5, Site(area_m2=100_000), norms=spb)
        assert res.density_chel_per_ga.status.value == "ok"

    def test_status_error_for_excessive_density(self, spb):
        """block_density=2.4 на 1 га — плотность превышает норму."""
        res = verify_kit(2.4, Site(area_m2=10_000), norms=spb)
        assert res.density_chel_per_ga.status.value == "error"

    def test_defaults_no_norms_arg(self):
        """Если norms не передан — загружает spb автоматически."""
        res = verify_kit(1.0, Site(area_m2=30_000))
        assert res.profile == "spb"

    def test_znop_threshold_uses_kit_developed(self, spb):
        """ЗНОП piecewise теперь смотрит на КИТ ПЗЗ (apt/lot), а не на
        block_density. На квартале 50 000 м² при разных block_density,
        kit_developed относительно стабилен и определяет ZNOP."""
        # Значения block_density: 0.8 и 2.0; разные kit_developed возможны,
        # но в обоих случаях ZNOP должен быть строго выводим из result.kit.
        res_low = verify_kit(0.8, Site(area_m2=50_000), norms=spb)
        res_high = verify_kit(2.0, Site(area_m2=50_000), norms=spb)
        # Проверяем что ZNOP согласован с КИТ ПЗЗ через piecewise:
        # piecewise breakpoints: 1.59→0, 1.79→3, 1.99→4, 2.50→6
        for r in (res_low, res_high):
            kit = r.kit.value
            expected = (
                0 if kit <= 1.59
                else 3 if kit <= 1.79
                else 4 if kit <= 1.99
                else 6
            )
            assert r.znop_per_person.value == expected, (
                f"при КИТ={kit:.3f} ожидался ZNOP={expected}, "
                f"получен {r.znop_per_person.value}"
            )


# ---------------------------------------------------------------------------
# Scenario
# ---------------------------------------------------------------------------

class TestScenario:
    def test_inverse_default(self):
        sc = Scenario(name="X", site=Site(area_m2=10_000))
        assert sc.mode == "inverse"
        assert sc.kit is None

    def test_verify_needs_kit(self):
        with pytest.raises(ValueError, match="kit обязателен"):
            Scenario(name="X", site=Site(area_m2=10_000), mode="verify")

    def test_verify_ok_with_kit(self):
        sc = Scenario(name="X", site=Site(area_m2=10_000), mode="verify", kit=1.5)
        assert sc.kit == 1.5


# ---------------------------------------------------------------------------
# run_scenarios / compare_scenarios
# ---------------------------------------------------------------------------

class TestRunScenarios:
    @pytest.fixture
    def scenarios(self):
        return [
            Scenario(name="Малый", site=Site(area_m2=20_000)),
            Scenario(
                name="Средний",
                site=Site(area_m2=50_000),
                options=CalculationOptions(floors=12, planning_doc=True),
            ),
            Scenario(
                name="Проверка КИТ=1.5",
                site=Site(area_m2=50_000),
                mode="verify",
                kit=1.5,
            ),
        ]

    def test_run_returns_correct_count(self, scenarios, spb):
        pairs = run_scenarios(scenarios, spb)
        assert len(pairs) == 3

    def test_verify_scenario_has_exact_block_density(self, scenarios, spb):
        """В режиме verify входной `kit` интерпретируется как block_density."""
        pairs = run_scenarios(scenarios, spb)
        name, res = pairs[2]
        assert name == "Проверка КИТ=1.5"
        assert res.block_density.value == 1.5

    def test_inverse_scenario_kit_is_pzz(self, scenarios, spb):
        """В inverse-режиме result.kit — это КИТ ПЗЗ; должен быть в разумных пределах."""
        pairs = run_scenarios(scenarios, spb)
        _, res = pairs[1]
        assert 0 < res.kit.value <= 3.0  # КИТ ПЗЗ может превышать 2.5 при infeasible

    def test_compare_returns_dataframe(self, scenarios, spb):
        df = compare_scenarios(scenarios, spb)
        assert "Малый" in df.columns
        assert "КИТ" in df.index

    def test_compare_columns_order(self, scenarios, spb):
        df = compare_scenarios(scenarios, spb)
        assert list(df.columns) == ["Малый", "Средний", "Проверка КИТ=1.5"]

    def test_no_ppt_max_kit_constrained(self, spb):
        """Без ДПТ КИТ_max=1.4 — норматив применяется к КИТ ПЗЗ. Если
        даже на минимальной плотности КИТ_ПЗЗ > 1.4, бисекция вернёт
        infeasible-результат: balance.is_feasible=False либо kit.status=ERROR.
        """
        site = Site(area_m2=200_000)
        scs = [
            Scenario(name="с ДПТ", site=site, options=CalculationOptions(planning_doc=True)),
            Scenario(name="без ДПТ", site=site, options=CalculationOptions(planning_doc=False)),
        ]
        pairs = run_scenarios(scs, spb)
        res_yes = pairs[0][1]
        res_no = pairs[1][1]
        # Без ДПТ: либо полностью OK (feasible баланс + KIT ≤ 1.4), либо
        # результат помечен как infeasible/error.
        # v0.9.16: balance.is_feasible может быть True даже при KIT > 1.4
        # (resurplus засчитан в зелень). Проверяем оба условия.
        from urban_model.models.result import Status
        fully_ok = (
            res_no.balance.is_feasible
            and res_no.kit.status != Status.ERROR
        )
        if fully_ok:
            assert res_no.kit.value <= 1.4 + 1e-3
        else:
            assert res_no.limiting_factor  # должен быть объяснён


# ---------------------------------------------------------------------------
# results_to_dataframe
# ---------------------------------------------------------------------------

class TestResultsToDataframe:
    def test_shape(self, spb, site_medium):
        pairs = [
            ("A", solve_max_kit(site_medium, norms=spb)),
            ("B", verify_kit(1.5, site_medium, norms=spb)),
        ]
        df = results_to_dataframe(pairs)
        assert "A" in df.columns
        assert "B" in df.columns
        assert "КИТ" in df.index
        assert "Статус баланса территории" in df.index


# ---------------------------------------------------------------------------
# xlsx-экспорт
# ---------------------------------------------------------------------------

class TestToXlsx:
    @pytest.fixture
    def pairs(self, spb):
        site = Site(area_m2=50_000)
        return [
            ("Обратный", solve_max_kit(site, norms=spb)),
            ("Проверка 1.8", verify_kit(1.8, site, norms=spb)),
        ]

    def test_creates_file(self, pairs):
        with tempfile.TemporaryDirectory() as tmp:
            path = pathlib.Path(tmp) / "report.xlsx"
            result_path = to_xlsx(pairs, path)
            assert result_path.exists()
            assert result_path.stat().st_size > 0

    def test_returns_path(self, pairs):
        with tempfile.TemporaryDirectory() as tmp:
            path = pathlib.Path(tmp) / "r.xlsx"
            result = to_xlsx(pairs, path)
            assert isinstance(result, pathlib.Path)

    def test_xlsx_has_correct_sheets(self, pairs):
        """Проверяем, что в книге есть листы Сравнение и Аудит."""
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as tmp:
            path = to_xlsx(pairs, pathlib.Path(tmp) / "r.xlsx")
            wb = load_workbook(path)
            assert "Сравнение" in wb.sheetnames
            assert "Аудит" in wb.sheetnames

    def test_comparison_sheet_has_data(self, pairs):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as tmp:
            path = to_xlsx(pairs, pathlib.Path(tmp) / "r.xlsx")
            wb = load_workbook(path)
            ws = wb["Сравнение"]
            # Первая строка — заголовок с именами сценариев
            header_vals = [c.value for c in ws[1]]
            assert "Обратный" in header_vals
            assert "Проверка 1.8" in header_vals
            # Не менее 10 строк данных
            assert ws.max_row >= 10
