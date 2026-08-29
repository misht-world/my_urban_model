"""Тесты v0.12.25 — комплексный «паспорт варианта» xlsx (Сводка + Баланс)."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from urban_model import solve_max_kit
from urban_model.export import build_variant_xlsx
from urban_model.models import CalculationOptions, Site
from urban_model.normatives import load_normatives


@pytest.fixture(scope="module")
def spb():
    return load_normatives("spb")


def _all_cells(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c is not None:
                out.append(str(c))
    return out


def test_two_sheets(spb, tmp_path):
    r = solve_max_kit(Site(area_m2=200_000), CalculationOptions(floors=16, planning_doc=True), spb)
    p = build_variant_xlsx("V1", r, CalculationOptions(floors=16, planning_doc=True), tmp_path / "v.xlsx")
    wb = load_workbook(p)
    assert wb.sheetnames == ["Сводка", "Баланс территории"]


def test_summary_has_all_categories(spb, tmp_path):
    opts = CalculationOptions(floors=16, planning_doc=True)
    r = solve_max_kit(Site(area_m2=300_000), opts, spb)
    p = build_variant_xlsx("V", r, opts, tmp_path / "v.xlsx")
    cells = _all_cells(load_workbook(p)["Сводка"])
    text = "\n".join(cells)
    for section in [
        "Основные показатели", "Жильё", "Социальные объекты", "Парковки",
        "ЗНОП и озеленение", "Проезды", "Инженерная инфраструктура",
        "Экономика (условные баллы)",
    ]:
        assert section in text, f"нет секции {section}"
    # доп. образование и инженерные объекты присутствуют
    assert any("Доп. образование" in c for c in cells)
    assert any("ТП" in c for c in cells)
    assert any("Эконом-индекс" in c for c in cells)


def test_balance_sheet_components(spb, tmp_path):
    opts = CalculationOptions(floors=16, planning_doc=True)
    r = solve_max_kit(Site(area_m2=300_000), opts, spb)
    p = build_variant_xlsx("V", r, opts, tmp_path / "v.xlsx")
    ws = load_workbook(p)["Баланс территории"]
    cells = _all_cells(ws)
    text = "\n".join(cells)
    assert "ЗУ жилой застройки" in text
    assert "Резерв (surplus)" in text
    assert "Контроль озеленения ТОП (6 м²/чел)" in text
    assert "Площадь квартала" in text


def test_no_economy_does_not_crash(spb, tmp_path):
    opts = CalculationOptions(floors=12, planning_doc=True, include_economy=False)
    r = solve_max_kit(Site(area_m2=150_000), opts, spb)
    p = build_variant_xlsx("V", r, opts, tmp_path / "v.xlsx")
    wb = load_workbook(p)
    assert "Сводка" in wb.sheetnames  # экономики нет, но файл собран
